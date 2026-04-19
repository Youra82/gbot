# src/gbot/analysis/show_results.py
"""
Analyse- und Anzeige-Skript fuer den gbot.

Modi:
  1) Einzel-Analyse            — jede Grid-Strategie wird isoliert getestet
  2) Manuelle Portfolio-Sim    — du waehlst das Team
  3) Auto Portfolio-Optimierung— der Bot waehlt das beste Team
  4) Interaktive Charts        — Fibonacci-Zonen + Grid-Levels
"""
import argparse
import glob
import json
import os
import sys
from datetime import date, datetime, timezone
from itertools import combinations

import pandas as pd

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR = os.path.join(PROJECT_ROOT, 'src', 'gbot', 'strategy', 'configs')
SETTINGS_FILE = os.path.join(PROJECT_ROOT, 'settings.json')


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def sep(char='=', width=68):
    print(char * width)


def load_configs() -> list:
    files = sorted(glob.glob(os.path.join(CONFIGS_DIR, 'config_*.json')))
    configs = []
    for path in files:
        try:
            with open(path) as f:
                configs.append((os.path.basename(path), json.load(f)))
        except Exception as e:
            print(f"  [WARNUNG] Config nicht lesbar ({os.path.basename(path)}): {e}")
    return configs


def _fetch_and_backtest(cfg: dict, start_date: str, end_date: str, capital: float) -> dict:
    """
    Laedt OHLCV-Daten und simuliert den Grid mit dynamischem Fibonacci-Rebalancing —
    identisch zum Live-Bot und zu Mode 4 (interaktive Charts).
    """
    from gbot.analysis.fibonacci import fetch_ohlcv_public
    from gbot.analysis.interactive_charts import simulate_dynamic_grid

    sym = cfg['market']['symbol']
    fib_cfg = cfg['grid'].get('fibonacci', {})
    tf = fib_cfg.get('timeframe', '4h')
    num_grids = cfg['grid']['num_grids']
    leverage = cfg['risk'].get('leverage', 1)
    lookback_fib = fib_cfg.get('lookback', 200)
    swing_window = fib_cfg.get('swing_window', 10)
    prefer_golden = fib_cfg.get('prefer_golden_zone', False)
    min_rebalance_h = fib_cfg.get('min_rebalance_interval_hours', 4)

    try:
        from gbot.analysis.optimizer import LOOKBACK_BY_TF, DEFAULT_LOOKBACK
        lookback_full = LOOKBACK_BY_TF.get(tf, DEFAULT_LOOKBACK)
        df = fetch_ohlcv_public(sym, tf, lookback_full)
        df = df[df.index >= start_date]
        df = df[df.index <= end_date]
        if len(df) < 10:
            return {'error': f'Zu wenige Kerzen ({len(df)}) im Zeitraum'}
    except Exception as e:
        return {'error': f'Datenabruf fehlgeschlagen: {e}'}

    try:
        grid_epochs, pnl_df, fills_df = simulate_dynamic_grid(
            df=df,
            num_grids=num_grids,
            leverage=leverage,
            capital=capital,
            lookback_fib=lookback_fib,
            swing_window=swing_window,
            prefer_golden_zone=prefer_golden,
            min_rebalance_hours=min_rebalance_h,
        )
    except Exception as e:
        return {'error': f'Grid-Simulation fehlgeschlagen: {e}'}

    if pnl_df.empty:
        return {'error': 'Keine PnL-Daten generiert'}

    total_pnl = float(pnl_df['pnl'].iloc[-1])
    roi_pct = total_pnl / capital * 100 if capital > 0 else 0
    total_fills = len(fills_df) if not fills_df.empty else 0

    cap_series = pnl_df['capital']
    running_max = cap_series.cummax()
    drawdown = (running_max - cap_series) / running_max * 100
    max_drawdown_pct = float(drawdown.max())

    first_epoch = grid_epochs[0] if grid_epochs else {}
    end_capital = round(capital + total_pnl, 4)
    return {
        'symbol': sym,
        'timeframe': tf,
        'num_grids': num_grids,
        'leverage': leverage,
        'total_fills': total_fills,
        'roi_pct': round(roi_pct, 2),
        'max_drawdown_pct': round(max_drawdown_pct, 2),
        'total_pnl_usdt': round(total_pnl, 4),
        'end_capital_usdt': end_capital,
        'lower': first_epoch.get('lower', 0),
        'upper': first_epoch.get('upper', 0),
        'lower_label': first_epoch.get('lower_label', 'n/a'),
        'upper_label': first_epoch.get('upper_label', 'n/a'),
        'candles': len(df),
        'n_rebalancings': len(grid_epochs) - 1 if grid_epochs else 0,
        'fib_analysis': {},
        'cap_series': pnl_df['capital'] if not pnl_df.empty else pd.Series(dtype=float),
        'fills_df': fills_df,
    }


# ---------------------------------------------------------------------------
# Portfolio-Hilfsfunktion: kombinierte Equity-Kurve & DD
# ---------------------------------------------------------------------------

def _combined_portfolio_dd(results: list, initial_capital: float) -> float:
    """
    Kombiniert N unabhaengige Equity-Kurven (jede startet bei initial_capital) zu einer
    Portfolio-Gesamtkurve (startet bei N * initial_capital) und berechnet den Max-DD
    davon. Gibt den korrekten Diversifikations-Vorteil wieder — Strategien die phasenverschoben
    sind, senken den Portfolio-DD unter das Maximum der Einzelwerte.
    """
    series_list = [r.get('cap_series') for r in results]
    valid = [s for s in series_list if s is not None and len(s) > 0]
    if not valid:
        return max(r['max_drawdown_pct'] for r in results)

    df = pd.concat(valid, axis=1).sort_index()
    df = df.ffill().bfill()

    portfolio = df.sum(axis=1)
    running_max = portfolio.cummax()
    dd = (running_max - portfolio) / running_max * 100
    return round(float(dd.max()), 2)


# ---------------------------------------------------------------------------
# Portfolio-Export: HTML-Chart + Excel (Mode 2 & 3)
# ---------------------------------------------------------------------------

def _get_telegram_cfg() -> tuple:
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json'), 'r') as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        return tg.get('bot_token', ''), tg.get('chat_id', '')
    except Exception:
        return '', ''


def _generate_gbot_chart(results: list, total_capital: float,
                          start_date: str, end_date: str, combined_dd: float):
    try:
        import plotly.graph_objects as go
    except ImportError:
        print("  plotly nicht installiert — Chart uebersprungen. (pip install plotly)")
        return

    n = len(results)
    COLORS = ['#f59e0b', '#10b981', '#8b5cf6', '#f97316', '#ec4899',
              '#14b8a6', '#a3e635', '#fb923c', '#e879f9', '#38bdf8']

    # Kombinierte Portfolio-Equity-Kurve
    series_list = [r['cap_series'] for r in results if r.get('cap_series') is not None and len(r.get('cap_series', [])) > 0]
    if not series_list:
        print("  Keine Equity-Daten — Chart uebersprungen.")
        return

    combined_df = pd.concat(series_list, axis=1).sort_index().ffill().bfill()
    portfolio = combined_df.sum(axis=1)

    total_pnl_pct = (portfolio.iloc[-1] - total_capital) / total_capital * 100
    sign = '+' if total_pnl_pct >= 0 else ''
    pairs = [f"{r['symbol'].split('/')[0]}/{r['timeframe']}" for r in results]
    title = (
        f"gbot Portfolio — {n} Strategie(n) ({', '.join(pairs)}) | "
        f"Zeitraum: {start_date} → {end_date} | "
        f"PnL: {sign}{total_pnl_pct:.1f}% | "
        f"Endkapital: {portfolio.iloc[-1]:.2f} USDT | "
        f"MaxDD: {combined_dd:.1f}%"
    )

    fig = go.Figure()
    fig.add_hline(y=total_capital,
                  line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
                  annotation_text=f'Start {total_capital:.0f} USDT',
                  annotation_position='top left')

    for idx, r in enumerate(results):
        cs = r.get('cap_series')
        if cs is None or len(cs) == 0:
            continue
        label = f"{r['symbol'].split('/')[0]}/{r['timeframe']}"
        fig.add_trace(go.Scatter(
            x=cs.index, y=cs.values,
            mode='lines', name=label,
            line=dict(color=COLORS[idx % len(COLORS)], width=1.2, dash='dot'),
            opacity=0.6,
            hovertemplate=f"{label}: %{{y:.2f}} USDT<extra></extra>",
        ))

    fig.add_trace(go.Scatter(
        x=portfolio.index, y=portfolio.values,
        mode='lines', name='Portfolio Equity',
        line=dict(color='#2563eb', width=2.5),
        hovertemplate='Portfolio: %{y:.2f} USDT<extra></extra>',
    ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=11), x=0.5, xanchor='center'),
        height=600, hovermode='x unified', template='plotly_dark',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        margin=dict(l=60, r=60, t=80, b=40),
        yaxis=dict(title='Equity (USDT)', fixedrange=False),
    )

    out_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'gbot_portfolio_equity.html')
    fig.write_html(out_file)
    print(f"  Chart gespeichert: gbot_portfolio_equity.html")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        try:
            from gbot.utils.telegram import send_document
            caption = (
                f"gbot Portfolio-Equity\n"
                f"{start_date} → {end_date} | {n} Strategie(n) | "
                f"PnL: {sign}{total_pnl_pct:.1f}% | "
                f"Equity: {portfolio.iloc[-1]:.2f} USDT | "
                f"MaxDD: {combined_dd:.1f}%"
            )
            send_document(bot_token, chat_id, out_file, caption=caption)
            print(f"  Chart via Telegram gesendet.")
        except Exception as e:
            print(f"  Telegram-Versand fehlgeschlagen: {e}")


def _generate_gbot_excel(results: list, total_capital: float,
                          start_date: str, end_date: str):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl nicht installiert — Excel uebersprungen. (pip install openpyxl)")
        return

    # Alle Fills aus allen Strategien sammeln
    all_fills = []
    for r in results:
        fdf = r.get('fills_df')
        cs = r.get('cap_series')
        if fdf is None or fdf.empty:
            continue
        sym_short = r['symbol'].split('/')[0]
        tf = r['timeframe']
        # Kapital zum Fill-Zeitpunkt aus cap_series interpolieren
        for ts, row in fdf.iterrows():
            if cs is not None and not cs.empty:
                idx_pos = cs.index.searchsorted(ts, side='right') - 1
                cap_at = float(cs.iloc[max(0, idx_pos)]) if idx_pos >= 0 else total_capital
            else:
                cap_at = None
            all_fills.append({
                'Datum':    str(ts)[:16].replace('T', ' '),
                'Symbol':   sym_short,
                'TF':       tf,
                'Seite':    'KAUF' if row['side'] == 'buy' else 'VERKAUF',
                'Preis':    round(float(row['price']), 6),
                'Kapital':  round(cap_at, 4) if cap_at is not None else '',
            })

    if not all_fills:
        print("  Keine Fill-Daten — Excel uebersprungen.")
        return

    all_fills.sort(key=lambda x: x['Datum'])
    for i, f in enumerate(all_fills, 1):
        f['Nr'] = i

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grid Fills'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    buy_fill     = PatternFill('solid', fgColor='D6F4DC')
    sell_fill    = PatternFill('solid', fgColor='FAD7D7')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    headers = ['Nr', 'Datum', 'Symbol', 'TF', 'Seite', 'Preis', 'Kapital']
    col_widths = {'Nr': 5, 'Datum': 18, 'Symbol': 10, 'TF': 6,
                  'Seite': 10, 'Preis': 14, 'Kapital': 14}

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 12)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(all_fills, 2):
        fill = buy_fill if row['Seite'] == 'KAUF' else (sell_fill if r_idx % 2 == 0 else alt_fill)
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Preis', 'Kapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    # Zusammenfassung
    sr = len(all_fills) + 3
    total_fills = len(all_fills)
    buys  = sum(1 for f in all_fills if f['Seite'] == 'KAUF')
    sells = sum(1 for f in all_fills if f['Seite'] == 'VERKAUF')
    end_cap = results[0]['cap_series'].iloc[-1] if results and results[0].get('cap_series') is not None and len(results[0].get('cap_series', [])) > 0 else 0
    pnl_pct = (sum(r.get('total_pnl_usdt', 0) for r in results) / total_capital * 100) if total_capital else 0

    for label, value in [
        ('Grid Fills gesamt', total_fills),
        ('Kaeufe', buys), ('Verkaeufe', sells),
        ('PnL', f"{pnl_pct:+.1f}%"),
        ('Startkapital', f"{total_capital:.2f} USDT"),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    out_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'gbot_portfolio_fills.xlsx')
    wb.save(out_file)
    print(f"  Excel gespeichert: gbot_portfolio_fills.xlsx  ({total_fills} Fills)")

    bot_token, chat_id = _get_telegram_cfg()
    if bot_token and chat_id:
        try:
            from gbot.utils.telegram import send_document
            caption = (
                f"gbot Grid Fills — {total_fills} Fills | "
                f"PnL: {pnl_pct:+.1f}%"
            )
            send_document(bot_token, chat_id, out_file, caption=caption)
            print(f"  Excel via Telegram gesendet.")
        except Exception as e:
            print(f"  Telegram-Versand fehlgeschlagen: {e}")


# ---------------------------------------------------------------------------
# Modus 1: Einzel-Analyse
# ---------------------------------------------------------------------------

def run_single_analysis(start_date: str, end_date: str, capital: float):
    sep()
    print("  gbot Einzel-Analyse — jede Strategie isoliert getestet")
    sep()
    print(f"  Zeitraum: {start_date} bis {end_date} | Gesamtkapital: {capital} USDT\n")

    configs = load_configs()
    if not configs:
        print("  Keine Config-Dateien gefunden. Bitte ./run_pipeline.sh ausfuehren.")
        return

    all_results = []
    for filename, cfg in configs:
        sym = cfg['market']['symbol']
        print(f"  Analysiere: {sym} ...", end=' ', flush=True)
        r = _fetch_and_backtest(cfg, start_date, end_date, capital)
        if r.get('error'):
            print(f"FEHLER: {r['error']}")
            continue
        print(f"OK ({r['candles']} Kerzen)")
        all_results.append({
            'Strategie': f"{sym} ({r['timeframe']})",
            'Grids': r['num_grids'],
            'Hebel': f"{r['leverage']}x",
            'Fills': r['total_fills'],
            'Reb.': r.get('n_rebalancings', 0),
            'ROI %': r['roi_pct'],
            'Max DD %': r['max_drawdown_pct'],
            'PnL USDT': r['total_pnl_usdt'],
            'Endkap. USDT': r['end_capital_usdt'],
        })

    if not all_results:
        print("\n  Keine gueltigen Ergebnisse.")
        return

    df = pd.DataFrame(all_results).sort_values('ROI %', ascending=False)
    pd.set_option('display.width', 140)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.float_format', '{:.2f}'.format)
    sep()
    print(df.to_string(index=False))
    sep()


# ---------------------------------------------------------------------------
# Modus 2: Manuelle Portfolio-Simulation
# ---------------------------------------------------------------------------

def run_manual_portfolio(start_date: str, end_date: str, capital: float):
    sep()
    print("  gbot Manuelle Portfolio-Simulation")
    sep()

    configs = load_configs()
    if not configs:
        print("  Keine Config-Dateien gefunden.")
        return

    print("\n  Verfuegbare Strategien:")
    for i, (filename, cfg) in enumerate(configs):
        sym = cfg['market']['symbol']
        tf = cfg['grid'].get('fibonacci', {}).get('timeframe', '?')
        print(f"    {i+1}) {sym} ({tf})")

    selection = input("\n  Welche Strategien simulieren? (Zahlen mit Komma, z.B. 1,3 oder 'alle'): ").strip()
    if selection.lower() == 'alle':
        selected = configs
    else:
        try:
            indices = [int(x.strip()) - 1 for x in selection.split(',')]
            selected = [configs[i] for i in indices]
        except (ValueError, IndexError):
            print("  Ungueltige Auswahl.")
            return

    print(f"\n  Zeitraum: {start_date} bis {end_date} | Gesamtkapital: {capital} USDT\n")

    results = []
    total_capital = capital * len(selected)
    total_pnl = 0.0
    total_fills = 0

    for filename, cfg in selected:
        sym = cfg['market']['symbol']
        print(f"  Backtest: {sym} ...", end=' ', flush=True)
        r = _fetch_and_backtest(cfg, start_date, end_date, capital)
        if r.get('error'):
            print(f"FEHLER: {r['error']}")
            continue
        print(f"ROI: {r['roi_pct']:+.2f}%")
        results.append(r)
        total_pnl += r['total_pnl_usdt']
        total_fills += r['total_fills']

    if not results:
        print("\n  Keine gueltigen Ergebnisse.")
        return

    portfolio_roi = (total_pnl / total_capital * 100) if total_capital > 0 else 0
    combined_dd = _combined_portfolio_dd(results, capital)
    max_dd_individual = max(r['max_drawdown_pct'] for r in results)

    sep()
    print("  Portfolio-Simulations-Ergebnis  (paralleler Live-Betrieb)")
    sep('-')
    for r in results:
        pnl_str = f"{r['total_pnl_usdt']:+.4f}"
        end_cap = r['end_capital_usdt']
        print(f"  {r['symbol']:<22} ({r['timeframe']})  ROI: {r['roi_pct']:>+7.2f}%  DD: {r['max_drawdown_pct']:>5.2f}%  PnL: {pnl_str} USDT  Endkap: {end_cap:.2f} USDT  Fills: {r['total_fills']}")
    sep('-')
    end_total = total_capital + total_pnl
    print(f"  Startkapital     : {total_capital:.2f} USDT  ({len(results)}x {capital:.2f} USDT parallel)")
    print(f"  Endkapital       : {end_total:.2f} USDT")
    print(f"  Gesamt-PnL       : {total_pnl:+.4f} USDT")
    print(f"  Portfolio ROI    : {portfolio_roi:+.2f}%")
    print(f"  Portfolio Max DD : {combined_dd:.2f}%  (Einzel-Max: {max_dd_individual:.2f}%)")
    print(f"  Gesamt-Fills     : {total_fills}")
    sep()

    print()
    ans = input("  Chart & Excel erstellen und via Telegram senden? (j/n) [Standard: n]: ").strip().lower()
    if ans in ('j', 'y', 'ja'):
        _generate_gbot_chart(results, total_capital, start_date, end_date, combined_dd)
        _generate_gbot_excel(results, total_capital, start_date, end_date)


# ---------------------------------------------------------------------------
# Modus 3: Automatische Portfolio-Optimierung
# ---------------------------------------------------------------------------

def run_auto_portfolio(start_date: str, end_date: str, capital: float, target_max_dd: float):
    sep()
    print("  gbot Automatische Portfolio-Optimierung")
    sep()
    print(f"  Ziel: Maximaler ROI bei maximal {target_max_dd:.1f}% Drawdown pro Grid.")
    print(f"  Zeitraum: {start_date} bis {end_date} | Gesamtkapital: {capital} USDT\n")

    configs = load_configs()
    if not configs:
        print("  Keine Config-Dateien gefunden.")
        return

    # Schritt 1: Alle Configs einzeln backtesten
    individual = []
    for filename, cfg in configs:
        sym = cfg['market']['symbol']
        print(f"  Backtest: {sym} ...", end=' ', flush=True)
        r = _fetch_and_backtest(cfg, start_date, end_date, capital)
        if r.get('error'):
            print(f"FEHLER: {r['error']}")
            continue
        print(f"ROI: {r['roi_pct']:+.2f}%  DD: {r['max_drawdown_pct']:.2f}%")
        individual.append((filename, cfg, r))

    if not individual:
        print("\n  Keine gueltigen Ergebnisse.")
        return

    # Schritt 2: Beste Kombination finden (bis 5er-Teams)
    best_roi = -9999.0
    best_team = []
    best_stats = {}

    max_team_size = min(5, len(individual))
    total_combos = sum(len(list(combinations(individual, k))) for k in range(1, max_team_size + 1))
    print(f"\n  Pruefe {total_combos:,} Portfolio-Kombinationen...\n")

    checked = 0
    PROGRESS_INTERVAL = max(1, total_combos // 200)

    for size in range(1, max_team_size + 1):
        for combo in combinations(individual, size):
            checked += 1
            results_combo = [r for _, _, r in combo]
            if any(r['roi_pct'] <= -9999 for r in results_combo):
                continue
            combined_dd = _combined_portfolio_dd(results_combo, capital)
            if combined_dd > target_max_dd:
                continue
            total_pnl = sum(r['total_pnl_usdt'] for r in results_combo)
            total_cap = capital * size
            roi = total_pnl / total_cap * 100
            if roi > best_roi:
                best_roi = roi
                best_team = combo
                best_stats = {
                    'total_pnl': total_pnl,
                    'total_capital': total_cap,
                    'roi': roi,
                    'combined_dd': combined_dd,
                    'max_dd_individual': max(r['max_drawdown_pct'] for r in results_combo),
                    'fills': sum(r['total_fills'] for r in results_combo),
                }

            if checked % PROGRESS_INTERVAL == 0 or checked == total_combos:
                pct = checked / total_combos * 100
                if best_team:
                    syms = ' + '.join(f"{r['symbol'].split('/')[0]}/{r['timeframe']}" for _, _, r in best_team)
                    best_str = f"Bestes: {syms}  ROI {best_roi:+.1f}%  DD {best_stats['combined_dd']:.1f}%"
                else:
                    best_str = "kein gueltiges Team bisher"
                print(f"\r  [{pct:5.1f}%] {checked:>{len(str(total_combos))}}/{total_combos}  —  {best_str:<60}", end='', flush=True)

    print()  # Zeilenumbruch nach Fortschrittsanzeige

    sep()
    if not best_team:
        print(f"  Kein Portfolio gefunden, das Max Drawdown <= {target_max_dd:.1f}% erfuellt.")
        print("  Versuche einen hoeher Drawdown-Wert oder mehr Trials.")
        sep()
        return

    n_team = len(best_team)
    print(f"  Optimales Team gefunden ({n_team} Strategie(n), {n_team}x {capital:.2f} USDT parallel):")
    sep('-')
    for filename, cfg, r in best_team:
        sym = r['symbol']
        end_cap = r['end_capital_usdt']
        print(f"  {sym:<22} ({r['timeframe']})  ROI: {r['roi_pct']:>+7.2f}%  DD: {r['max_drawdown_pct']:>5.2f}%  Endkap: {end_cap:.2f} USDT  Fills: {r['total_fills']}")
    sep('-')
    end_total = best_stats['total_capital'] + best_stats['total_pnl']
    print(f"  Startkapital     : {best_stats['total_capital']:.2f} USDT  ({n_team}x {capital:.2f} USDT)")
    print(f"  Endkapital       : {end_total:.2f} USDT")
    print(f"  Gesamt-PnL       : {best_stats['total_pnl']:+.4f} USDT")
    print(f"  Portfolio ROI    : {best_stats['roi']:+.2f}%")
    print(f"  Portfolio Max DD : {best_stats['combined_dd']:.2f}%  (Einzel-Max: {best_stats['max_dd_individual']:.2f}%)")
    print(f"  Gesamt-Fills     : {best_stats['fills']}")
    sep()

    # Ergebnis speichern fuer Shell-Nachbearbeitung
    serializable_stats = {k: v for k, v in best_stats.items()}
    opt_result = {
        'optimal_portfolio': [filename for filename, _, _ in best_team],
        'stats': serializable_stats,
        'generated_at': datetime.now(timezone.utc).isoformat(),
    }
    results_dir = os.path.join(PROJECT_ROOT, 'artifacts', 'results')
    os.makedirs(results_dir, exist_ok=True)
    out_path = os.path.join(results_dir, 'optimization_results.json')
    with open(out_path, 'w') as f:
        json.dump(opt_result, f, indent=2)
    print(f"  Ergebnis gespeichert: {out_path}")

    best_results = [r for _, _, r in best_team]
    _generate_gbot_chart(best_results, best_stats['total_capital'],
                          start_date, end_date, best_stats['combined_dd'])
    _generate_gbot_excel(best_results, best_stats['total_capital'],
                          start_date, end_date)


# ---------------------------------------------------------------------------
# Modus 4: Interaktive Charts (Plotly HTML)
# ---------------------------------------------------------------------------

def run_interactive_charts():
    try:
        from gbot.analysis.interactive_charts import main as charts_main
        charts_main()
    except Exception as e:
        print(f"\n  Fehler beim Ausfuehren der interaktiven Charts: {e}")
        import traceback
        traceback.print_exc()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="gbot Analyse-Tool")
    parser.add_argument('--mode', type=str, default='1', choices=['1', '2', '3', '4'])
    parser.add_argument('--start_date', type=str, default=None)
    parser.add_argument('--end_date', type=str, default=None)
    parser.add_argument('--capital', type=float, default=None)
    parser.add_argument('--target_max_drawdown', type=float, default=30.0)
    args = parser.parse_args()

    mode = args.mode

    # Modus 4 hat eigenes Eingabe-System
    if mode == '4':
        run_interactive_charts()
        return

    # Datum und Kapital abfragen
    start_date = args.start_date or input("  Startdatum (JJJJ-MM-TT) [Standard: 2023-01-01]: ").strip() or "2023-01-01"
    end_date = args.end_date or input(f"  Enddatum   (JJJJ-MM-TT) [Standard: Heute]:      ").strip() or date.today().strftime("%Y-%m-%d")
    cap_input = args.capital
    if cap_input is None:
        raw = input("  Gesamtkapital in USDT          [Standard: 100]:   ").strip()
        cap_input = float(raw) if raw else 100.0

    print()

    if mode == '1':
        run_single_analysis(start_date, end_date, cap_input)
    elif mode == '2':
        run_manual_portfolio(start_date, end_date, cap_input)
    elif mode == '3':
        run_auto_portfolio(start_date, end_date, cap_input, args.target_max_drawdown)


if __name__ == '__main__':
    main()
